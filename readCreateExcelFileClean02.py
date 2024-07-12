from openpyxl import load_workbook

from datetime import date
today = date.today()

# load the excel
wb = load_workbook(filename = "new_Clients_Savings_Group_2024-05-21.xlsx")

# grab the active Sheet in the excel
activeWorkSheet = wb.active

# Create a new WorkBook to store the manipulated data
from openpyxl import Workbook

# Create a new workbook, the final file
newWorkbookExcel = Workbook()

# Create a new workSheet in the new WorkBook
newWorkbookExcel.create_sheet('CopiedFromNewClientsExcel')
newWorkbookWorksheet01Excel = newWorkbookExcel['CopiedFromNewClientsExcel']

# Obtain Max Row and Max Column of the sheet in the raw file new_Clients_Savings_Group_2024-05-21.xlsx
sourceMaxRowInSheet = activeWorkSheet.max_row
sourceMaxColumnInSheet = activeWorkSheet.max_column

# copying the cell values to newWorkbookExcel from new_Clients_Savings_Group_2024-05-21.xlsx
for i in range(1, sourceMaxRowInSheet + 1):
    for j in range(1, sourceMaxColumnInSheet + 1):
        # reading cell value from newWorkbookExcel
        c = activeWorkSheet.cell(row=i, column=j)
        # writing the read value to newWorkbookExcel
        newWorkbookWorksheet01Excel.cell(row=i, column=j).value = c.value

# Create an empty list to store the data from column A
column_a_data = []

# Iterate through each row in column A and append the value to the list
for row in activeWorkSheet['A2:A100']:
    for cell in row:
        if cell.value is not None:
            column_a_data.append(cell.value)
        else:
            # Handle empty cells, as per your requirements
            column_a_data.append(None)  # Placeholder value or None, based on your preference

# Create an empty list to store the data from column C
column_c_data = []

for row in activeWorkSheet['C2:C100']:
    for cell in row:
        if cell.value is not None:
            column_c_data.append(cell.value)
        else:
            column_c_data.append(None)  # Placeholder value or None, based on your preference

# debugging
#print("column_a_data: ",column_a_data)
#print()
#print("column_c_data: ",column_c_data)

zipColumnAColumnC = zip(column_c_data,column_a_data)
listZip = list(zipColumnAColumnC)
# debugging
#print("listZip: ",listZip)

# grouping via Dictionaries
from collections import defaultdict

grouped_data_02 = defaultdict(list)
for key, value in listZip:
    grouped_data_02[key].append(value)

# debugging
#print("dict(grouped_data_02): ",dict(grouped_data_02))

newListGroup = []
newListPortfolios = []
for key, value in dict(grouped_data_02).items():
    #print(f'{key}: {value}')
    newListGroup.append(key)
    newListPortfolios.append(value)

# debugging
#print("newListGroup \n:", newListGroup)
#print("newListPortfolios \n:", newListPortfolios)
#print("type(newListPortfolios)",type(newListPortfolios))
#print("break line \n")

columnFList = []
concatenate = ""
for eachItem in newListPortfolios:
    #print("eachItem: ",eachItem)
    for eachSubItem in eachItem:
        #print("eachSubItem: ",eachSubItem)
        #print(type(eachSubItem))
        concatenate = str(eachSubItem) + ";" + concatenate
    #print("concatenate: ", concatenate)
    columnFList.append(concatenate)
    #print("columnFList: ", columnFList)
    concatenate = ""
    #print()

# debugging
#print("columnFList: ",columnFList)

# to save the above 2 lists; column_e_data, column_f_data, and new WorkSheet into new excel

# Create a new workSheet in the new WorkBook
newWorkbookExcel.create_sheet('GroupingDone')
newWorkbookWorksheet02Excel = newWorkbookExcel['GroupingDone']

# transfer list to column
for index, itemInCellColumnF in enumerate(columnFList):
    newIndex = index + 1
    newWorkbookWorksheet02Excel.cell(row=newIndex + 1, column=6, value=itemInCellColumnF)

# add headers to excel
newWorkbookWorksheet02Excel.cell(row=1, column=6).value="List of Portfolios per Grouping"

unique_values_columnEList = []
for item in column_c_data:
    if item not in unique_values_columnEList:
        unique_values_columnEList.append(item)
#print("unique_values_columnEList: ",unique_values_columnEList)

# transfer list to column
for index, itemInCellColumnE in enumerate(unique_values_columnEList):
    newIndex = index + 1
    newWorkbookWorksheet02Excel.cell(row=newIndex + 1, column=5, value=itemInCellColumnE)

# add headers to excel
newWorkbookWorksheet02Excel.cell(row=1, column=5).value="Unique Block/Group"

# Save the workbook --> create the new excel
newWorkbookExcel.save('new_Clients_Savings_Group_' + str(today) + '.xlsx')

# Close the old Excel file
wb.close()

# Close the new Excel file
newWorkbookExcel.close()